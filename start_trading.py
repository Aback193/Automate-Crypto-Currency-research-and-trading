import time
import indicators
import build_model
import binance_trade

import joblib
import numpy as np
import pandas as pd
import datetime
from datetime import datetime as dtn
from pandas.core.frame import DataFrame
import pandas_datareader.data as web
import yfinance as yf
import matplotlib.pyplot as plt
import csv
import pydotplus
import mplfinance as mpf
import os
from openpyxl import Workbook, load_workbook
from binance.client import Client   # pip install python-binance

# Binance API key goes here
api = {0: {'api_key' : 'API_KEY', 'api_secret' : 'API_KEY'}}
client = Client(api[0]['api_key'], api[0]['api_secret'])

classifier = 'load' # = retrain to retrain model if exists; = load just to laod existent one;
subfolderModels = "./BinanceModels/"
startDate = datetime.datetime(2007,1,1)
endDate = datetime.datetime.now()
USDT_symbols = []
targetGain = 50     # % gain target
indicatorsToUse = ['RSI','STOCH','Williams','MACD','ROC','OBV','ADX','CMF']   # ['RSI','STOCH','Williams','MACD','ROC','OBV','ADX'] # ['RSI','STOCH','Williams'] highest weight, add 'CMF' for crypto but not for Stocks
coin = 'ETH'
tickerBinance = coin + 'USDT'
ticker = coin + '-USD'
#ticker = 'ETH-USD'  # GOOGL # AAPL # TSLA # S&P = ^GSPC # BTC-USD # ETH-USD # SOL1-USD # XTZ-USD
timeframe = '15m'

total_sentiment = {}

def open_workbook(path, sheet_name):
    workbook = load_workbook(filename=path)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        #print(f'The title of the Worksheet is: {sheet.title}')
        row_len = int(sheet.calculate_dimension().split(':')[1].split('D')[1])
        for i in range(2, row_len+1):
            #print(f'The value of B is {sheet["B"+str(i)].value}')
            #print(f'The value of C is {sheet["C"+str(i)].value}')
            if sheet["B"+str(i)].value in total_sentiment:
                total_sentiment[sheet["B"+str(i)].value] = total_sentiment[sheet["B"+str(i)].value] + float(sheet["C"+str(i)].value)
            else:
                total_sentiment[sheet["B"+str(i)].value] = float(sheet["C"+str(i)].value)
        #print(f"Cells that contain data: {sheet.calculate_dimension()}")
        #print(f"Cells that contain data: {row_len}")
        for i, (k, v) in enumerate(total_sentiment.items()):
            #print("DICT num {}".format(i))
            #print("DICT key {}".format(k))
            total_sentiment[k] = float(v) / float((row_len-1)/total_sentiment.__len__())
            #print("DICT value {}".format(total_sentiment[k]))

def getBinanceHistoricalData(tickerPair, tf, startDate):
    try:
        # Get market depth, bids and asks on Binance 
        #depth = client.get_order_book(symbol=tickerPair)
        #depth_df = pd.DataFrame(depth['asks'])      # depth_df = pd.DataFrame(depth['bids'])
        #depth_df.columns = ['Price', 'Volume']
        #depth_df.head()

        # Get Historical Data

        #  [
        #   [
        #     1499040000000,      // Open time
        #     "0.01634790",       // Open
        #     "0.80000000",       // High
        #     "0.01575800",       // Low
        #     "0.01577100",       // Close
        #     "148976.11427815",  // Volume
        #     1499644799999,      // Close time
        #     "2434.19055334",    // Quote asset volume
        #     308,                // Number of trades
        #     "1756.87402397",    // Taker buy base asset volume
        #     "28.46694368",      // Taker buy quote asset volume
        #     "17928899.62484339" // Ignore.
        #   ]
        # ]

        historical = client.get_historical_klines(tickerPair, tf, startDate)
        hist_df = pd.DataFrame(historical)
        hist_df.columns = ['Open Time', 'Open', 'High', 'Low', 'Close', 'Volume', 'Close Time', 'Quote Asset Volume', 
                    'Number of Trades', 'TB Base Volume', 'TB Quote Volume', 'Ignore']
        # Preprocess Historical Data, becuse dtypes are Strings and Int's
        hist_df['Open Time'] = pd.to_datetime(hist_df['Open Time']/1000, unit='s')      # COnvert to human readable Date Time Format
        hist_df['Close Time'] = pd.to_datetime(hist_df['Close Time']/1000, unit='s')      # COnvert to human readable Date Time Format
        numeric_columns = ['Open', 'High', 'Low', 'Close', 'Volume', 'Quote Asset Volume', 'TB Base Volume', 'TB Quote Volume']
        hist_df[numeric_columns] = hist_df[numeric_columns].apply(pd.to_numeric, axis=1)        # Convert String columns to Float64        
        return hist_df
    except Exception as e:
        print("\ngetBinanceHistoricalData Exception: {}".format(e))
        return DataFrame()

def getInitialHistoricalData():
    print('\nFetching Binance history for {} on {} timeframe'.format(tickerBinance, timeframe))
    df_temp = getBinanceHistoricalData(tickerBinance, timeframe, '1 Jan 2011')
    #df_temp = yf.download(tickers=ticker, period='60d', interval=timeframe)    # Valid intervals: [1m, 2m, 5m, 15m, 30m, 60m, 90m, 1h, 1d, 5d, 1wk, 1mo, 3mo]
    while len(df_temp) == 0:
        time.sleep(60)
        df_temp = getBinanceHistoricalData(tickerBinance, timeframe, '1 Jan 2011')
        #df_temp = yf.download(tickers=ticker, period='60d', interval=timeframe)

    #print(df_temp.dtypes)    # Print var types in DataFrame by Column
    #print(df_temp.tail())
    df_temp.to_csv(subfolderModels+'history_binance_'+tickerBinance+'_'+timeframe+'.csv')
    #candleStickChart(df_temp)
    return df_temp

# save pre trained model to file
def save_model(model):
    joblib.dump(model, subfolderModels+"random_forest_"+ticker+"_"+timeframe+".joblib")

# Calculate indicatos with fresh data from Yahoo or Binance !!!! NEED TO MAKE IT LAST 20 CANDLES BASED ON TF !!!!
def freshIndicatorCalculations():
    try:
        todayTemp = datetime.date.today()
        yesterdayTemp = todayTemp - datetime.timedelta(days=21)
        twoDaysAgo = yesterdayTemp.strftime('%d %b %Y')
        print(twoDaysAgo)
        #df_try = yf.download(tickers=ticker, period='1d', interval=timeframe)
        df_try = getBinanceHistoricalData(tickerBinance, timeframe, twoDaysAgo)
        print('\nFetching Binance history for {} on {} timeframe for last 2 days'.format(tickerBinance, timeframe))
        if not len(df_try) == 0:
            df_try['RSI'] = indicators.rsi(df_try)
            df_try['STOCH'] = indicators.stoch(df_try)
            df_try['Williams'] = indicators.williams(df_try)
            df_try['MACD'] = indicators.macd(df_try)
            df_try['ROC'] = indicators.roc(df_try)
            df_try['OBV'] = indicators.obv(df_try)
            df_try['ADX'] = indicators.adx(df_try)
            df_try['CMF'] = indicators.cmf(df_try)
            indicatorsLastRow = df_try.iloc[-1]
            selection = pd.DataFrame(indicatorsLastRow[indicatorsToUse]).transpose()
            #print("Indicators last row Binance: " + str(selection.values))
            #df_try.to_csv(subfolderModels+'fresh_indicators.csv')
            #selection.to_csv(subfolderModels+'indicators_last_row.csv')
            return selection.values
        else:
            return DataFrame()
    except Exception as e:
        print("\ndef freshIndicatorCalculations EXCEPTION: " + str(e))
        return DataFrame()

# Execute Binance Buy or Sell based on prediction
def livePredicitonsBinance(freshIndicators):

    ticker_key = ticker.split('-')[0]
    predict = model.predict(freshIndicators)
    print(ticker + " PREDICTIONS: {}".format(predict[0]))
    
    if str(predict[0]) == '-1.0' and total_sentiment[ticker_key] < 0.4:
        print("\nSELL, waiting for binance_trade")
        binance_trade.init("SELL", tickerBinance)
        #print("\nSELL, waiting no buy no sell")
    elif str(predict[0]) == '1.0' and total_sentiment[ticker_key] > 0.7:
        print("\nBUY, waiting for binance_trade")
        binance_trade.init("BUY", tickerBinance)
    elif str(predict[0]) == '0.0' and (total_sentiment[ticker_key] > 0.4 and total_sentiment[ticker_key] < 0.7):
        print("\nNEUTRAL, waiting no buy no sell")


if __name__ == "__main__":

    open_workbook("../Sentiment.xlsx", "Total sentiment")

    # Create a new directory if it does not exist
    if not os.path.exists(subfolderModels) and subfolderModels != "./":
        os.makedirs(subfolderModels)

    ################################### Train save and load saved model ###################################
    if classifier == "retrain":
        try:
            df = getInitialHistoricalData()
            print("\nBuilding model...\n")
            # Train model and save it to file
            save_model(build_model.buildRandomForestModel(df, tickerBinance, timeframe))
            #df.to_csv(subfolderModels+'decision_tree_table'+tickerBinance+'.csv')
            # load random forest pre trained model from file
            model = joblib.load(subfolderModels+"random_forest_"+ticker+"_"+timeframe+".joblib")
        except Exception as e:
            print("\nRetrain of the model Exception: {}".format(e))
    else:
        try:
            model = joblib.load(subfolderModels+"random_forest_"+ticker+"_"+timeframe+".joblib")
        except Exception as e:
            print("\nTrained model load Exception: {}".format(e))
            df = getInitialHistoricalData()
            print("\nBuilding model...\n")
            save_model(build_model.buildRandomForestModel(df, tickerBinance, timeframe))
            #df.to_csv(subfolderModels+'decision_tree_table'+tickerBinance+'.csv')
            model = joblib.load(subfolderModels+"random_forest_"+ticker+"_"+timeframe+".joblib")
    ################################### Train save and load saved model ###################################

    ################################### Live Predictions ###################################
    while True:
        now = dtn.now()
        print("\nWaiting for the end of the current candle...")
        if timeframe == "15m":
            minutes = ["14", "29", "44", "59"]
            current_time_min = now.strftime("%M")
            current_time_sec = now.strftime("%S")
            while not (any(x == current_time_min for x in minutes) and current_time_sec == "01"):
                now = dtn.now()
                current_time_min = now.strftime("%M")
                current_time_sec = now.strftime("%S")
        elif timeframe == "30m":
            current_time_min = now.strftime("%M")
            while not (current_time_min == "29" or current_time_min == "59"):
                now = dtn.now()
                current_time_min = now.strftime("%M")
        elif timeframe == "1h":
            current_time_min = now.strftime("%M")
            while not current_time_min == "59":
                now = dtn.now()
                current_time_min = now.strftime("%M")
        elif timeframe == "1d":
            current_time_hour = now.strftime("%H")
            current_time_min = now.strftime("%M")
            while not (current_time_hour == "00" and current_time_min == "59"):
                now = dtn.now()
                current_time_hour = now.strftime("%H")
                current_time_min = now.strftime("%M")
        #freshIndicators = ta_api.getIndicatorValue('ETH/USDT', timeframe)
        freshIndicators = freshIndicatorCalculations()
        while len(freshIndicators) == 0:
            print("\nfreshIndicators DF is Empty, Sleeping 60s")
            time.sleep(60)
            #freshIndicators = ta_api.getIndicatorValue('ETH/USDT', timeframe)
            freshIndicators = freshIndicatorCalculations()
        if not len(freshIndicators) == 0:
            print("\n" + ticker + " Fresh indicator values: {}".format(freshIndicators))
            livePredicitonsBinance(freshIndicators)
        print("\nSleeping 16s")
        time.sleep(16)
    ################################### Live Predictions ###################################

